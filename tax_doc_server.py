import os
import re
import openpyxl
from pypdf import PdfReader
from fastmcp import FastMCP

mcp = FastMCP("ITR_Production_Engine")
SECURE_DOCS_DIR = os.path.expanduser("~/Documents/Financials/ITR-2026/")


def sanitize_text(text: str) -> str:
    """Hardened local regex sanitization path for Indian tax identifiers."""
    if not text:
        return ""
    text = re.sub(r'\b[A-Za-z]{5}[\s-]*[0-9]{4}[\s-]*[A-Za-z]\b', '[REDACTED_PAN]', text)
    text = re.sub(r'\b\d{4}[\s-]*\d{4}[\s-]*\d{4}\b', '[REDACTED_AADHAAR]', text)
    text = re.sub(r'\b\d{9,18}\b', '[REDACTED_ACCOUNT_OR_TX_ID]', text)
    text = re.sub(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', '[REDACTED_EMAIL]', text)
    text = re.sub(r'\b(?:\+91[\s-]?)?[6-9]\d{9}\b', '[REDACTED_PHONE]', text)
    return text


def parse_to_flat_lines(filename: str) -> list:
    """Converts PDF, Excel, or Text structures into clean array segments securely."""
    file_path = os.path.join(SECURE_DOCS_DIR, filename)
    lines = []
    ext = filename.lower().split('.')[-1]

    if ext == 'pdf':
        reader = PdfReader(file_path)
        for idx, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                lines.append(f"--- BLOCK START: PAGE {idx + 1} ---")
                lines.extend(text.split('\n'))
    elif ext in ['xlsx', 'xls']:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"--- BLOCK START: SHEET {sheet} ---")
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    lines.append(" | ".join(str(c) if c is not None else "" for c in row))
    else:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [line.strip() for line in f.readlines()]
    return lines


@mcp.tool()
def list_tax_directory() -> str:
    """Lists all processable document filenames inside the secure ITR folder."""
    if not os.path.exists(SECURE_DOCS_DIR):
        return f"Directory {SECURE_DOCS_DIR} does not exist yet."

    files = os.listdir(SECURE_DOCS_DIR)
    supported_files = [f for f in files if f.lower().endswith(('.pdf', '.xlsx', '.xls', '.csv', '.txt'))]

    if not supported_files:
        return "The repository is empty or contains no supported financial documents."

    return "Available Documents:\n" + "\n".join(f"- {f}" for f in supported_files)


@mcp.tool()
def inspect_tax_document(filename: str) -> str:
    """
    Analyzes document structures. Returns full data if small,
    or a line-indexed map if it's a massive ledger.
    """
    try:
        raw_lines = parse_to_flat_lines(filename)
        total_lines = len(raw_lines)

        # Small document auto-stream path
        if total_lines <= 600:
            return sanitize_text("\n".join(raw_lines))

        # Large dataset index mapper
        summary = [
            f"=== LARGE FILE SUMMARY: {filename} ===",
            f"Total Lines Detected: {total_lines}",
            "To target specific sections, utilize the 'read_document_chunk' tool.",
            "\n--- LINE INDEX MAP OF HEADERS & DATA FIELDS ---"
        ]

        for index, line in enumerate(raw_lines):
            if "BLOCK START" in line or any(k in line.upper() for k in
                                            ["DATE", "PAN", "SUMMARY", "TOTAL", "TAX SCHEDULE", "CAPITAL GAINS",
                                             "BALANCE", "STATEMENT"]):
                summary.append(f"Line {index:04d}: {sanitize_text(line[:90])}")

        return "\n".join(summary)
    except Exception as e:
        return f"Error evaluating data schema for {filename}: {str(e)}"


@mcp.tool()
def read_document_chunk(filename: str, start_line: int, end_line: int) -> str:
    """Extracts a precise bounding box of lines from an indexed large document."""
    try:
        raw_lines = parse_to_flat_lines(filename)
        total_lines = len(raw_lines)

        start = max(0, start_line)
        end = min(total_lines, end_line)

        if start >= total_lines:
            return f"Error: Requested line index bounds out of scope. Total file size is {total_lines} lines."

        chunk = raw_lines[start:end]
        header = f"=== PAYLOAD CHUNK: Lines {start} to {end} of {total_lines} ===\n"

        return header + sanitize_text("\n".join(chunk))
    except Exception as e:
        return f"Error slicing document frame: {str(e)}"


if __name__ == "__main__":
    mcp.run(transport="stdio")